"""Teste do parser: detecção do cabeçalho por âncora num .xlsx em memória.

Reproduz a estrutura real: o cabeçalho 'Data|Conta|...' aparece DUAS vezes; só a
segunda (seguida por datas de verdade) é a tabela. Há painel lateral e rodapé sem data.
"""

import io
import datetime as dt

import openpyxl

from ingestion.contabil.parser import (
    parse_xlsx, COL_DATA, COL_CONTA, COL_TIPO, COL_SETOR,
    COL_CATEGORIA, COL_PROJETO, COL_VALOR, COL_OBS,
)


def _planilha_fake() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    # Helper: escreve numa linha (1-based) a partir da coluna D (4, 1-based)
    def linha(r, data, conta, tipo, setor, cat, proj, valor, obs):
        ws.cell(r, COL_DATA + 1, data)
        ws.cell(r, COL_CONTA + 1, conta)
        ws.cell(r, COL_TIPO + 1, tipo)
        ws.cell(r, COL_SETOR + 1, setor)
        ws.cell(r, COL_CATEGORIA + 1, cat)
        ws.cell(r, COL_PROJETO + 1, proj)
        ws.cell(r, COL_VALOR + 1, valor)
        ws.cell(r, COL_OBS + 1, obs)

    # Painel lateral / título (ruído nas colunas A-B)
    ws.cell(1, 2, "VISÃO GERAL")
    # PRIMEIRO cabeçalho (mini-tabela), NÃO seguido por datas
    hdr = ["Data", "Conta", "Entrada/Saída", "Setor", "Categoria",
           "Nº do Projeto", "Valor", "Observações"]
    for j, h in enumerate(hdr):
        ws.cell(5, COL_DATA + 1 + j, h)
    # (linha 6 fica vazia -> esse cabeçalho é descartado)

    # SEGUNDO cabeçalho (tabela real), seguido de datas
    for j, h in enumerate(hdr):
        ws.cell(10, COL_DATA + 1 + j, h)
    linha(11, dt.datetime(2026, 6, 1), "Cora", "Saída", "Operações", "Clicksign", "", 229.62, "Clicksign")
    linha(12, dt.datetime(2026, 5, 31), "Cora", "Entrada", "Presidência", "ENEJ", "010.2026", 257.25, "Pacote")
    # rodapé sem data -> ignorado
    ws.cell(13, COL_DATA + 1, "TOTAL")
    ws.cell(14, COL_DATA + 1, dt.time(0, 0))  # time, não date -> ignorado

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parser_acha_segunda_tabela_e_ignora_ruido():
    linhas = parse_xlsx(_planilha_fake())
    assert len(linhas) == 2                      # só as 2 com data; rodapé/time fora
    assert linhas[0]["conta"] == "Cora"
    assert linhas[0]["valor"] == 229.62
    assert linhas[0]["data"] == dt.date(2026, 6, 1)
    assert linhas[1]["projeto"] == "010.2026"
    assert linhas[0]["row_num"] == 11


def test_parser_aceita_bytes():
    assert isinstance(_planilha_fake(), bytes)
    assert parse_xlsx(_planilha_fake())          # não levanta
